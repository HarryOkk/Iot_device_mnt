from remes_aliyun_openapi.iot.device_manage import batch_query_device_detail
from remes_aliyun_openapi.iot.device_manage import query_device_prop
from remes_aliyun_openapi.iot.thing_model_use import set_devices_property
from remes_aliyun_openapi.iot.thing_model_use import query_device_desired_property
from remes_aliyun_openapi.iot.thing_model_use import query_device_property_status
import pandas as pd
import numpy as np
import openpyxl
import logging
import json
import typing
from remes_mysql.db_config import AliyunBizDb

SMEC_GW = 'g4xdEqa2CfX'
SMEC_CPD = 'g4xdsqZciZ0'
iot_instance_id = "iot-060a02m5"
excel_file_info = {'sheet_name': 'Sheet3',
                   'file_path': r"C:\Users\10010010\Desktop\上海中心.xlsx",
                   'cpdid_column_id': 0,
                   'dev_status_column_id': 1,
                   'device_thing_version_column': 2,
                   'cpd_cid_bind_column': 3,
                   'rgw_lable_column': 4,
                   'LIC_ele_contract_column': 5,
                   'receiver_id': 6,
                   'receiver_name': 7,
                   'customer_name': 8,
                   'ele_org_name': 9,
                   'ele_branch_org_name': 10,
                   'mnt_project_name': 11,
                   'mnt_project_address': 12,
                   'mnt_ele_org_name': 13,
                   'mnt_ele_branch_org_name': 14,
                   'ele_local_name': 15,
                   'mnt_contract_id': 20
                   }

# 日志模块初始化
device_mnt_logger = logging.Logger('阿里设备脚本方式管理日志')
device_mnt_handler = logging.FileHandler(r'D:\software\Notepad++\change.log')
device_mnt_handler.setLevel(logging.DEBUG)
device_mnt_format = logging.Formatter(
    '文件：%(filename)s  时间：%(asctime)s  函数：%(funcName)s  级别：%(levelname)s \n 内容：%(message)s')
device_mnt_handler.setFormatter(device_mnt_format)
device_mnt_logger.addHandler(device_mnt_handler)


# 获取excel_file_info['file_path']的excel表格中装置的列表
def extract_colum_cpdid() -> list:
    xlsx = pd.ExcelFile(excel_file_info['file_path'])
    # 该函数导出xlsx文件中的所有sheet
    df = pd.read_excel(xlsx, excel_file_info['sheet_name'])
    try:
        l_devices = df.iloc[:, excel_file_info['cpdid_column_id']].tolist()
    except:
        l_devices = []
        device_mnt_logger.info(f"导出CPDID列表失败，问题列表为：{excel_file_info['sheet_name']}")
    device_mnt_logger.info(f"从文件{excel_file_info['file_path']}中拿到的所有控制柜装置如下：{l_devices}")
    device_mnt_logger.info(f'装置的一共有{len(l_devices)}台')
    return l_devices


# 获取excel_file_info['file_path']的excel表格中电梯的列表
def extract_colum_cid_eles() -> list:
    xlsx = pd.ExcelFile(excel_file_info['file_path'])
    # 该函数导出xlsx文件中的所有sheet
    df = pd.read_excel(xlsx, excel_file_info['sheet_name'])
    try:
        l_eles = df.iloc[:, excel_file_info['LIC_ele_contract_column']].tolist()
    except:
        l_eles = []
        device_mnt_logger.info(f"导出电梯列表失败，问题列表为：{excel_file_info['sheet_name']}")
    device_mnt_logger.info(f"从文件{excel_file_info['file_path']}中拿到的电梯如下：{l_eles}")
    device_mnt_logger.info(f'电梯的一共有{len(l_eles)}台')
    return l_eles


# 将表格中的电梯合同号梯号列返回成python列表
def extract_ele_contract() -> list:
    xlsx = pd.ExcelFile(excel_file_info['file_path'])
    df = pd.read_excel(xlsx, excel_file_info['sheet_name'])
    try:
        l_contract_eles = df.iloc[:, excel_file_info['LIC_ele_contract_column']].tolist()

    except:
        l_contract_eles = []
        device_mnt_logger.info(f"导出CID列表失败：{excel_file_info['sheet_name']}")

    device_mnt_logger.info(f"从文件{excel_file_info['file_path']}中拿到的所有cid如下：{l_contract_eles}")
    device_mnt_logger.info(f'可追溯的电梯的一共有{len(l_contract_eles)}台')
    return l_contract_eles


# 查询指定设备的当前物模型属性值
def query_devices_thing_status(function_block_id: str,
                               thing_lable_name: str,
                               l_devices: list[str],
                               excel_dev_line: int,
                               excel_thing_module_line: int
                               ) -> dict:
    d_response = {}
    for device_name in l_devices:
        if function_block_id == 'default':
            d_func_block = query_device_property_status(iot_instance_id=iot_instance_id,
                                                        product_key=SMEC_CPD,
                                                        device_name=device_name)
        else:
            d_func_block = query_device_property_status(iot_instance_id=iot_instance_id,
                                                        product_key=SMEC_CPD,
                                                        device_name=device_name,
                                                        function_block_id=function_block_id)
        l_thing_of_device = d_func_block['body']['Data']['List']['PropertyStatusInfo']
        for d_single_func_thing in l_thing_of_device:
            try:
                if d_single_func_thing['Identifier'] == thing_lable_name:
                    d_response[device_name] = d_single_func_thing['Value']
            except KeyError:
                d_response[device_name] = None
    for key, word in d_response.items():
        write_to_excel_by_match(match_line_of_excel=excel_dev_line,
                                writen_line_of_excel=excel_thing_module_line,
                                match_item=key,
                                writen_item=word
                                )
    return d_response


# 查询装置的标签字典
def query_devices_label(l_devices: list[int]) -> dict:
    dict_devices_labels = {}
    for device in l_devices:
        response = query_device_prop(iot_instance_id, SMEC_GW, str(device))
        device_mnt_logger.info('下面是装置的标签信息')
        device_mnt_logger.info(response)
        if response['body']['Success']:
            try:
                dev_props = response['body']['Props']
                dict_dev_labels = json.loads(dev_props)
                dict_devices_labels[str(device)] = dict_dev_labels
            except KeyError:
                device_mnt_logger.info(f'设备{device}不包含标签')
                print(f'设备{device}不包含标签')
        else:
            device_mnt_logger.info(device)
            device_mnt_logger.info('装置调用接口失败')
            device_mnt_logger.info(response['body'])
    return dict_devices_labels


# 无线通话装置中，将RCPD上传的RGW标签写入到excel表格中
def write_rgw_lable_to_excel(l_devices: list[str]):
    d_label = query_devices_label(l_devices)
    print(d_label)
    for device_name_cpd, d_device_cpd_label in d_label.items():
        try:
            rgw = d_device_cpd_label['REMOTE_GW']
            writen_result = write_to_excel_by_match(match_line_of_excel=excel_file_info['cpdid_column_id'],
                                                    writen_line_of_excel=excel_file_info['rgw_lable_column'],
                                                    match_item=device_name_cpd,
                                                    writen_item=rgw)
        except KeyError:
            print(f'{device_name_cpd}装置标签写入结果不含有rgw标签')


# 查询装置列表的在线状态，并将在线状态填写到表格对应的位置
def query_devices_status_to_xlsx(l_devices: list[str]):
    wb = openpyxl.load_workbook(excel_file_info['file_path'])
    # 获取表中第一个sheet对象,用于写入表格
    sheet = wb[excel_file_info['sheet_name']]
    chunk_size = 100  # 分片大小,因该接口每次调用最多查询100台装置，所以需要对list进行分片
    for i in range(0, len(l_devices), chunk_size):
        chunk = l_devices[i:i + chunk_size]
        # 在这里对每个分片进行处理

        response_dev_mnt = batch_query_device_detail(SMEC_GW, chunk, iot_instance_id)
        device_mnt_logger.info(response_dev_mnt)
        if response_dev_mnt['body']['Success']:
            # 从返回值中拿到设备详细信息的列表
            l_device_detail = response_dev_mnt['body']['Data']['Data']
            device_mnt_logger.info(l_device_detail)
            print('表格中装置的状态为：')
            for dict_dev_item in l_device_detail:
                print(dict_dev_item['DeviceName'], dict_dev_item['Status'])
                # 将装置的在线状态更新到表格中
                for row in sheet.iter_rows():
                    if str(row[excel_file_info['cpdid_column_id']].value) == dict_dev_item['DeviceName']:
                        row_num = row[excel_file_info['cpdid_column_id']].row
                        # 把装置的在线状态写入对应的表格
                        row_later = sheet[row_num]
                        row_later[excel_file_info['dev_status_column_id']].value = dict_dev_item['Status']
        else:
            print('装置在线状态查询接口调用失败，详细信息为：', response_dev_mnt['body'])
    wb.save(excel_file_info['file_path'])


# 开通装置列表种所有装置的CRT参数
def activate_param_crt(l_devices: list[str]) -> dict:
    items = '{"pCRT": 1}'
    response = set_devices_property(SMEC_GW, items, l_devices, iot_instance_id)
    print("开通CRT参数调用结果为：", response['body'])
    return response


# 开通列表所有装置的LIC参数
def activate_param_license(l_devices: list) -> dict:
    items = '{"WirelessCallFunc": 1}'
    response = set_devices_property(SMEC_GW, items, l_devices, iot_instance_id)
    print("开通无线通话参数参数调用结果为：")
    print(response['body'])
    return response


# 根据列表中装置的cpd_id，在无线通话开通表中查询对应的合同号梯号(注意：传入的cpdid列表元素应该为str类型)
def query_license_device(l_device: list[str]) -> bool:
    dict_lic_device_cid = {}
    for device_name_cpd in l_device:
        sql = f'''
            select
                ele_id
            from
                license_cpd_elevator
            where
                cpd_id = {device_name_cpd}
        '''
        df_cpd_ele_id = AliyunBizDb().read_data(sql=sql)
        print(df_cpd_ele_id, type(df_cpd_ele_id))
        try:
            current_ele_id = df_cpd_ele_id["ele_id"][0]
            sql = f'''
                select
                    ele_contract_no
                from
                    remes_elevator_base
                where
                    ele_id = '{current_ele_id}'
            '''
            df_cid_ele_id = AliyunBizDb().read_data(sql=sql)
            print(df_cid_ele_id, type(df_cid_ele_id))
            dict_lic_device_cid[device_name_cpd] = df_cid_ele_id["ele_contract_no"][0]
            write_to_excel_by_match(match_line_of_excel=excel_file_info['cpdid_column_id'],
                                    writen_line_of_excel=excel_file_info['LIC_ele_contract_column'],
                                    match_item=device_name_cpd,
                                    writen_item=df_cid_ele_id["ele_contract_no"][0]
                                    )
        except IndexError:
            print(f'{device_name_cpd}查询失败，该装置未通过申请LIC绑定电梯')
    return True


# 通过cpdid查询对应的合同号梯号
def query_cid_from_cpdid(cpdid: int):
    sql = f"""
            select
                ele_contract_no,
                cpd_id
            from
                t_cpd_elevator 
            inner join 
                remes_elevator_base 
            on
                t_cpd_elevator.ele_id=remes_elevator_base.ele_id
            where
               cpd_id={cpdid}
       """
    df = AliyunBizDb().read_data(sql=sql)
    print(df)
    if len(df) > 0:
        return df['ele_contract_no'][0]
    else:
        return None
        device_mnt_logger.info(f'在绑定关系表中未找到控制柜装置{cpdid}的绑定记录')


# 通过cid查询对应二维码绑定关系的CPDID
def query_cpdid_from_cid(cid: str):
    sql = f"""
            select
                ele_contract_no,
                cpd_id
            from
                t_cpd_elevator 
            inner join 
                remes_elevator_base 
            on
                t_cpd_elevator.ele_id=remes_elevator_base.ele_id
            where
               ele_contract_no='{cid}'
       """
    print(sql)
    df = AliyunBizDb().read_data(sql=sql)
    print(df)
    if len(df) > 0:
        return df['cpd_id'][0]
    else:
        return None
        device_mnt_logger.info(f'在绑定关系表中未找到电梯{cid}的绑定记录')


# 查询电梯列表（cid）的绑定关系并打印到filepath中
def query_bind_relationship_ele_and_write_to_excel(l_cid_eles: list):
    wb = openpyxl.load_workbook(excel_file_info['file_path'])
    # 获取表中第一个sheet对象,用于写入表格
    sheet = wb[excel_file_info['sheet_name']]

    for ele in l_cid_eles:
        query_ele_response = query_cpdid_from_cid(ele)
        for row in sheet.iter_rows():
            if str(row[excel_file_info['LIC_ele_contract_column']].value) == str(ele):
                row_num = row[excel_file_info['LIC_ele_contract_column']].row
                row_later = sheet[row_num]
                if query_ele_response is not None:
                    row_later[excel_file_info['cpd_cid_bind_column']].value = '二维码已绑定'
                    row_later[excel_file_info['cpdid_column_id']].value = query_ele_response
                else:
                    row_later[excel_file_info['cpd_cid_bind_column']].value = '二维码未绑定'
        wb.save(excel_file_info['file_path'])


# 查询对应的物模型期望值版本
def write_desire_thing_version_to_excel(
        l_devices: list,
        thing_identifier: str
):
    identifier = 'WirelessCallFunc'
    wb = openpyxl.load_workbook(excel_file_info['file_path'])
    # 获取表中第一个sheet对象,用于写入表格
    sheet = wb.active
    for device_name in l_devices:
        deivce_thing_version = query_device_desired_property(
            iot_instance_id=iot_instance_id,
            product_key=SMEC_GW,
            identifier=thing_identifier,
            device_name=device_name)
        dict_writed_desire_thing = {}
        for row in sheet.iter_rows():
            if str(row[excel_file_info['cpdid_column_id']].value) == str(device_name):
                row_num = row[excel_file_info['cpdid_column_id']].row
                # 把装置的在线状态写入对应的表格
                row_later = sheet[row_num]
                if deivce_thing_version is not None:
                    dict_WirelessCallFunc = \
                        [d for d in deivce_thing_version['body']['Data']['List']['DesiredPropertyInfo'] if
                         d['Identifier'] == 'WirelessCallFunc'][0]
                    print(type(dict_WirelessCallFunc), dict_WirelessCallFunc)
                    dict_writed_desire_thing[dict_WirelessCallFunc['Identifier']] = dict_WirelessCallFunc['Value']
                    row_later[excel_file_info['device_thing_version_column']].value = str(dict_writed_desire_thing)
                else:
                    row_later[excel_file_info['device_thing_version_column']].value = '未查到期望值版本'
        wb.save(excel_file_info['file_path'])


# 指定表格的列，找到匹配项，并将内容写入表格
def write_to_excel_by_match(match_line_of_excel: int,
                            writen_line_of_excel: int,
                            match_item: str,
                            writen_item: str
                            ) -> bool:
    wb = openpyxl.load_workbook(excel_file_info['file_path'])
    # 获取表中一个sheet对象,用于写入表格
    sheet = wb[excel_file_info['sheet_name']]

    for row in sheet.iter_rows():
        if str(row[match_line_of_excel].value) == str(match_item):
            row_num = row[match_line_of_excel].row
            # 把经过匹配后的信息写入表格
            row_later = sheet[row_num]
            row_later[writen_line_of_excel].value = writen_item
    wb.save(excel_file_info['file_path'])
    return True


# 查询装置列表（cpdid）的绑定关系并打印到filepath中
def query_bind_relationship_cpdid_and_write_to_excel(l_devices: list):
    d_binds = {}
    for device in l_devices:
        sql = f"""
            select
                cpd_id,
                ele_id,
                status
            from
                t_cpd_elevator 

            where
               cpd_id='{device}'        
        """
        df = AliyunBizDb().read_data(sql=sql)
        if len(df) > 0:
            write_to_excel_by_match(
                match_line_of_excel=excel_file_info['cpdid_column_id'],
                writen_line_of_excel=excel_file_info['cpd_cid_bind_column'],
                match_item=df['cpd_id'][0],
                writen_item=df['status'][0])


# 查询装置的REMES领用人
def query_device_receiver_to_excel(l_devices: list[str]) -> bool:
    dict_device_receiver_cid = {}
    for device_name_cpd in l_devices:
        try:
            sql = f'''
                select
                    device_id
                from
                    remes_device_base
                where
                    vender_terminal_no = {device_name_cpd}
                    '''
            df_device_id = AliyunBizDb().read_data(sql=sql)
            current_device_id = df_device_id["device_id"][0]
            sql = f'''
                    select
                        receiver_id, receiver_name
                    from
                        remes_device_repertory
                    where
                        device_id = '{current_device_id}'
                '''
            df_device_receiver_info = AliyunBizDb().read_data(sql=sql)
            print(df_device_receiver_info)
            write_to_excel_by_match(match_line_of_excel=excel_file_info['cpdid_column_id'],
                                    writen_line_of_excel=excel_file_info['receiver_id'],
                                    match_item=device_name_cpd,
                                    writen_item=df_device_receiver_info["receiver_id"][0])
            write_to_excel_by_match(match_line_of_excel=excel_file_info['cpdid_column_id'],
                                    writen_line_of_excel=excel_file_info['receiver_name'],
                                    match_item=device_name_cpd,
                                    writen_item=df_device_receiver_info["receiver_name"][0]
                                    )

        except IndexError:
            print(f'{device_name_cpd}查询失败，检查excel是否有领用人的列，若有，该装置未发放')
    return True


# 综合二维码绑定和无线声网开通两个信息来查询装置所绑定的电梯合同号梯号
def query_ele_mnt_info_to_excel(l_devices: list[str]) -> bool:
    dict_device_receiver_cid = {}
    for device_name_cpd in l_devices:
        sql = f'''
            select
                device_id
            from
                remes_device_base
            where
                vender_terminal_no = {device_name_cpd}
                '''
        df_device_id = AliyunBizDb().read_data(sql=sql)
        print(df_device_id, type(df_device_id))
        try:
            current_device_id = df_device_id["device_id"][0]
            sql = f'''
                select
                    receiver_id, receiver_name
                from
                    remes_device_repertory
                where
                    device_id = '{current_device_id}'
            '''
            df_device_receiver_info = AliyunBizDb().read_data(sql=sql)
            print(df_device_receiver_info, type(df_device_receiver_info))
            write_to_excel_by_match(match_line_of_excel=excel_file_info['cpdid_column_id'],
                                    writen_line_of_excel=excel_file_info['receiver_id'],
                                    match_item=device_name_cpd,
                                    writen_item=df_device_receiver_info["receiver_id"][0])
            write_to_excel_by_match(match_line_of_excel=excel_file_info['cpdid_column_id'],
                                    writen_line_of_excel=excel_file_info['receiver_name'],
                                    match_item=device_name_cpd,
                                    writen_item=df_device_receiver_info["receiver_name"][0]
                                    )
        except IndexError:
            print(f'{device_name_cpd}查询失败，该装置未发放')
    return True


# 通过查询数据库remes_elevator_base，达到根据合同号梯号查询各种维保信息的目的
def query_mnt_info_from_cid_to_excel() -> bool:
    l_eles_cid = extract_ele_contract()
    print(l_eles_cid)
    l_mnt_columns = ['customer_name', 'ele_org_name', 'ele_branch_org_name', 'mnt_project_name', 'mnt_project_address',
                     'mnt_ele_org_name', 'mnt_ele_branch_org_name', 'mnt_contract_id', 'ele_local_name']
    for ele in l_eles_cid:
        print(ele, type(ele))
        if ele != np.nan:
            sql = f'''
                select
                    reb.customer_name,
                    reb.ele_org_name,
                    reb.ele_branch_org_name,
                    reb.mnt_project_name,
                    reb.mnt_project_address,
                    reb.mnt_ele_org_name,
                    reb.mnt_ele_branch_org_name,
                    reb.mnt_contract_id,
                    reb2.ele_local_name 
                from
                    remes_elevator_base reb
                left join 
                    remes_elevator_business reb2 
                on 
                    reb.ele_id = reb2.ele_id
                where
                    ele_contract_no = '{ele}'
                    '''
            df_mnt_info = AliyunBizDb().read_data(sql=sql)
            if len(df_mnt_info) > 0:
                for mnt_name in l_mnt_columns:
                    write_to_excel_by_match(match_line_of_excel=excel_file_info['LIC_ele_contract_column'],
                                            writen_line_of_excel=excel_file_info[f'{mnt_name}'],
                                            match_item=ele,
                                            writen_item=df_mnt_info[f'{mnt_name}'][0]
                                            )
            else:
                pass
    return True


# 将维保信息按照excel_file_info（）写入到对应的表格中
def make_cpd_mnt_info_to_excel(l_devices: list[str]) -> bool:
    if l_devices:
        # 无线转换装置下控制柜装置rgw标签
        write_rgw_lable_to_excel(l_devices)
        # 查询装置在线状态
        query_devices_status_to_xlsx(l_devices)
        # 查询装置二维码绑定关系
        query_bind_relationship_cpdid_and_write_to_excel(l_devices)
        # 查询装置领用人信息
        print('是否完成查询装置领用人：', query_device_receiver_to_excel(l_devices))
        # 查询装置无线通话绑定的电梯情况
        print('无线通话绑定电梯查询结果为：', query_license_device(l_devices))
        # 查询装置通过合同号梯号获取的维保信息
        print('合同号梯号维保信息查询的结果为：', query_mnt_info_from_cid_to_excel())

    else:
        print(f"表格{excel_file_info['file_path']}不含有装置")
        return False
    return True


# 根据表格中的合同号梯号，查询维保ID，并填入表格中
def query_mnt_contract_id_by_cid() -> bool:
    l_eles_cid = extract_ele_contract()
    print(l_eles_cid)
    mnt_columns = 'mnt_contract_id'
    for ele in l_eles_cid:
        print(ele, type(ele))
        if ele != np.nan:
            sql = f'''
                       select
                           mnt_contract_id
                       from
                           remes_elevator_base
                       where
                           ele_contract_no = '{ele}'
                           '''
            df_mnt_info = AliyunBizDb().read_data(sql=sql)
            if len(df_mnt_info) > 0:
                write_to_excel_by_match(match_line_of_excel=excel_file_info['LIC_ele_contract_column'],
                                        writen_line_of_excel=excel_file_info[mnt_columns],
                                        match_item=ele,
                                        writen_item=df_mnt_info[mnt_columns][0]
                                        )
            else:
                pass
    return True


if __name__ == "__main__":
    # l_cid_eles = extract_colum_cid_eles()
    # query_bind_relationship_cpdid_and_write_to_excel(l_cid_eles)
    l_devices_from_excel = extract_colum_cpdid()
    print('维保信息查询写入结果为：', make_cpd_mnt_info_to_excel(l_devices_from_excel))
    # print("电梯楼层列表的查询结果为", query_devices_thing_status(
    #     function_block_id='default',
    #     thing_lable_name='EleDispFloorTable',
    #     l_devices=l_devices_from_excel,
    #     excel_dev_line=0,
    #     excel_thing_module_line=17
    # ))
