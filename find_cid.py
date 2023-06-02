from remes_mysql.db_config import AliyunBizDb
import pandas as pd
import openpyxl

file_path1 = r"D:\my_documents\智慧电梯\无线通话问题排查\季师傅需求\无线通话装置编号查询.xlsx"
file_path2 = r"D:\my_documents\智慧电梯\无线通话问题排查\季师傅需求\无线通话已绑定.xlsx"


def get_ele_cno(device_name):
    sql = f"""
        select
            ele_contract_no,
            vender_terminal_no
        from
            md_elevator_v
        where
            vender_terminal_no='{device_name}'
    """
    df = AliyunBizDb().read_data(sql=sql)
    if len(df) > 0:
        return df["ele_contract_no"][0]
    else:
        return None


if __name__ == "__main__":
    # 查询合同号后填入到表格
    # df = pd.read_excel(r"D:\my_documents\智慧电梯\新顺花苑1.xlsx")
    # df["ele_contract_no"] = df["监控装置DTU编号"].apply(lambda x: get_ele_cno(x))
    # df.to_excel("合同号梯号.xlsx")
    wb1 = openpyxl.load_workbook(file_path1)
    wb2 = openpyxl.load_workbook(file_path2)
    sheet1 = wb1.active
    sheet2 = wb2.active
    for CurrentRow2 in sheet2.iter_rows():
        for CurrentRow1 in sheet1.iter_rows():
            if CurrentRow2[1].value == CurrentRow1[2].value:
                CurrentRow1[3].value = CurrentRow2[0].value
            else:
                print(CurrentRow2[1].value, '不是绑定装置')
    wb1.save(file_path1)
    wb2.save(file_path2)
